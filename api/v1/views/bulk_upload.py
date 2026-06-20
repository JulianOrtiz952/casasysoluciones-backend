import re
from io import BytesIO
from django.db import transaction
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from rest_framework import status
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from pot.models import CustomUser, Property
from pot.services import property_service
from api.v1.permissions import IsAdmin
from api.v1.exceptions import APIError


# Helpers and Mappings
def _normalize_header(value):
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = re.sub(r'\s+', '_', text)
    text = (
        text.replace('á', 'a')
        .replace('é', 'e')
        .replace('í', 'i')
        .replace('ó', 'o')
        .replace('ú', 'u')
        .replace('ñ', 'n')
    )
    return text


def _cell_str(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


PROPERTY_HEADER_MAP = {
    'direccion': 'direccion',
    'address': 'direccion',
    'propietario': 'propietario',
    'owner_name': 'propietario',
    'dueno': 'propietario',
    'titulo': 'propietario',
    'title': 'propietario',
    'tipo': 'tipo',
    'type': 'tipo',
    'ciudad': 'ciudad',
    'city': 'ciudad',
    'edificio': 'edificio_conjunto',
    'conjunto': 'edificio_conjunto',
    'edificio_conjunto': 'edificio_conjunto',
    'unidad': 'unidad_apartamento',
    'apartamento': 'unidad_apartamento',
    'unidad_apartamento': 'unidad_apartamento',
    'estado': 'estado',
    'status': 'estado',
    'precio': 'precio',
    'price': 'precio',
    'habitaciones': 'habitaciones',
    'rooms': 'habitaciones',
    'banos': 'banos',
    'bathrooms': 'banos',
    'salas': 'salas',
    'living_rooms': 'salas',
    'cocinas': 'cocinas',
    'kitchens': 'cocinas',
    'garajes': 'garajes',
    'garages': 'garajes',
    'es_comercial': 'es_comercial',
    'es_comercial?_(s/n)': 'es_comercial',
    'comercial': 'es_comercial',
    'en_conjunto': 'en_conjunto',
    'en_conjunto?_(s/n)': 'en_conjunto',
    'administracion_incluida': 'administracion_incluida',
    'administracion_incluida?_(s/n)': 'administracion_incluida',
    'valor_administracion': 'valor_administracion',
    'admin_value': 'valor_administracion',
    'enlace_google_maps': 'enlace_google_maps',
    'google_maps_link': 'enlace_google_maps',
    'descripcion': 'descripcion',
    'description': 'descripcion',
    'observaciones': 'observaciones',
    'observations': 'observaciones',
}

USER_HEADER_MAP = {
    'correo_electronico': 'correo_electronico',
    'correo': 'correo_electronico',
    'email': 'correo_electronico',
    'nombres': 'nombres',
    'nombre': 'nombres',
    'first_name': 'nombres',
    'apellidos': 'apellidos',
    'apellido': 'apellidos',
    'last_name': 'apellidos',
    'tipo_documento': 'tipo_documento',
    'document_type': 'tipo_documento',
    'numero_documento': 'numero_documento',
    'documento': 'numero_documento',
    'document_number': 'numero_documento',
    'telefono': 'telefono',
    'phone': 'telefono',
    'rol': 'rol',
    'role': 'rol',
}

PROPERTY_TYPE_CHOICES_MAP = {
    'APARTAMENTO': Property.Type.APARTMENT,
    'CASA': Property.Type.HOUSE,
    'LOCAL': Property.Type.LOCAL,
    'BODEGA': Property.Type.WAREHOUSE,
    'APARTMENT': Property.Type.APARTMENT,
    'HOUSE': Property.Type.HOUSE,
    'LOCAL COMERCIAL': Property.Type.LOCAL,
    'WAREHOUSE': Property.Type.WAREHOUSE,
}

PROPERTY_STATUS_CHOICES_MAP = {
    'DISPONIBLE': Property.Status.AVAILABLE,
    'ARRENDADO': Property.Status.RENTED,
    'MANTENIMIENTO': Property.Status.MAINTENANCE,
    'AVAILABLE': Property.Status.AVAILABLE,
    'RENTED': Property.Status.RENTED,
    'MAINTENANCE': Property.Status.MAINTENANCE,
}

USER_ROLE_CHOICES_MAP = {
    'ARRENDATARIO': CustomUser.Role.TENANT,
    'INQUILINO': CustomUser.Role.TENANT,
    'ASISTENTE': CustomUser.Role.ASSISTANT,
    'ASISTENTE ADMINISTRATIVO': CustomUser.Role.ASSISTANT,
    'TECNICO': CustomUser.Role.TECHNICIAN,
    'ADMINISTRADOR': CustomUser.Role.ADMIN,
    'ADMIN': CustomUser.Role.ADMIN,
    'TENANT': CustomUser.Role.TENANT,
    'ASSISTANT': CustomUser.Role.ASSISTANT,
    'TECHNICIAN': CustomUser.Role.TECHNICIAN,
}

USER_DOC_TYPE_CHOICES_MAP = {
    'CC': CustomUser.DocumentType.CC,
    'CE': CustomUser.DocumentType.CE,
    'PASAPORTE': CustomUser.DocumentType.PASSPORT,
    'PASSPORT': CustomUser.DocumentType.PASSPORT,
    'NIT': CustomUser.DocumentType.NIT,
}


def _style_excel_sheet(ws, title):
    # Corporate style: Rose/Red color palette for header
    rose_fill = PatternFill(start_color="E11D48", end_color="E11D48", fill_type="solid") # Tailwind rose-600
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Format headers (Row 1)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = rose_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    # Auto-adjust columns width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    ws.row_dimensions[1].height = 28


# Views
class BulkUploadTemplateView(APIView):
    """Descarga de plantilla vacía para carga masiva de inmuebles o usuarios."""
    permission_classes = [IsAdmin]

    def get(self, request):
        upload_type = request.query_params.get('type')
        if upload_type not in ('properties', 'users'):
            return Response({'error': 'El parámetro "type" debe ser "properties" o "users".'}, status=status.HTTP_400_BAD_REQUEST)

        wb = Workbook()
        ws = wb.active

        if upload_type == 'properties':
            ws.title = "Plantilla Inmuebles"
            headers = [
                "Dirección*", "Título*", "Tipo (APARTAMENTO/CASA/LOCAL/BODEGA)*", 
                "Ciudad", "Edificio/Conjunto", "Unidad/Apartamento", 
                "Estado (DISPONIBLE/ARRENDADO/MANTENIMIENTO)", "Precio", 
                "Habitaciones", "Baños", "Salas", "Cocinas", "Garajes", 
                "Es Comercial? (S/N)", "En Conjunto? (S/N)", 
                "Administración Incluida? (S/N)", "Valor Administración", 
                "Enlace Google Maps", "Descripción", "Observaciones"
            ]
            ws.append(headers)
            # Add one example row
            ws.append([
                "Calle 123 # 45-67", "Juan Pérez", "APARTAMENTO", 
                "Bogotá", "Torres del Parque", "Apto 501", 
                "DISPONIBLE", 1500000, 3, 2, 1, 1, 1, 
                "N", "S", "S", 250000, 
                "https://maps.google.com/?q=Calle+123", "Hermoso apartamento familiar.", 
                "Ninguna observación."
            ])
            filename = "plantilla_inmuebles.xlsx"
        else:
            ws.title = "Plantilla Usuarios"
            headers = [
                "Correo Electrónico*", "Nombres*", "Apellidos*", 
                "Tipo Documento (CC/CE/PASAPORTE/NIT)*", "Número Documento*", 
                "Teléfono", "Rol (ARRENDATARIO/ASISTENTE/TECNICO/ADMINISTRADOR)"
            ]
            ws.append(headers)
            # Add one example row
            ws.append([
                "arrendatario.ejemplo@email.com", "Carlos", "Mendoza", 
                "CC", "1234567890", "3001234567", "ARRENDATARIO"
            ])
            filename = "plantilla_usuarios.xlsx"

        _style_excel_sheet(ws, ws.title)

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class BulkUploadExportView(APIView):
    """Exporta los datos existentes de inmuebles o usuarios utilizando el mismo formato de plantilla."""
    permission_classes = [IsAdmin]

    def get(self, request):
        upload_type = request.query_params.get('type')
        if upload_type not in ('properties', 'users'):
            return Response({'error': 'El parámetro "type" debe ser "properties" o "users".'}, status=status.HTTP_400_BAD_REQUEST)

        wb = Workbook()
        ws = wb.active

        if upload_type == 'properties':
            ws.title = "Inmuebles"
            headers = [
                "Dirección*", "Título*", "Tipo (APARTAMENTO/CASA/LOCAL/BODEGA)*", 
                "Ciudad", "Edificio/Conjunto", "Unidad/Apartamento", 
                "Estado (DISPONIBLE/ARRENDADO/MANTENIMIENTO)", "Precio", 
                "Habitaciones", "Baños", "Salas", "Cocinas", "Garajes", 
                "Es Comercial? (S/N)", "En Conjunto? (S/N)", 
                "Administración Incluida? (S/N)", "Valor Administración", 
                "Enlace Google Maps", "Descripción", "Observaciones"
            ]
            ws.append(headers)
            
            for p in Property.objects.all().order_by('-created_at'):
                ws.append([
                    p.address,
                    p.owner_name,
                    p.get_type_display().upper(),
                    p.city,
                    p.building_name,
                    p.unit_label,
                    p.get_status_display().upper(),
                    p.price,
                    p.rooms,
                    p.bathrooms,
                    p.living_rooms,
                    p.kitchens,
                    p.garages,
                    "S" if p.is_commercial else "N",
                    "S" if p.in_complex else "N",
                    "S" if p.admin_included else "N",
                    p.admin_value,
                    p.google_maps_link,
                    p.description,
                    p.observations
                ])
            filename = "exportacion_inmuebles.xlsx"
        else:
            ws.title = "Usuarios"
            headers = [
                "Correo Electrónico*", "Nombres*", "Apellidos*", 
                "Tipo Documento (CC/CE/PASAPORTE/NIT)*", "Número Documento*", 
                "Teléfono", "Rol (ARRENDATARIO/ASISTENTE/TECNICO/ADMINISTRADOR)"
            ]
            ws.append(headers)
            
            for u in CustomUser.objects.all().order_by('-created_at'):
                ws.append([
                    u.email,
                    u.first_name,
                    u.last_name,
                    u.document_type,
                    u.document_number or '',
                    u.phone,
                    u.get_role_display().upper()
                ])
            filename = "exportacion_usuarios.xlsx"

        _style_excel_sheet(ws, ws.title)

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        response = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class BulkUploadValidateView(APIView):
    """Parsea el archivo excel y corre validaciones en memoria sin escribir a base de datos."""
    permission_classes = [IsAdmin]
    parser_classes = [MultiPartParser]

    def post(self, request):
        upload_type = request.query_params.get('type')
        if upload_type not in ('properties', 'users'):
            raise APIError('invalid_type', 'El parámetro "type" debe ser "properties" o "users".', status_code=400)

        file_obj = request.FILES.get('file')
        if not file_obj:
            raise APIError('file_required', 'Debe enviar el archivo en el campo "file".', status_code=400)

        if not file_obj.name.lower().endswith('.xlsx'):
            raise APIError('invalid_file_type', 'Solo se admiten archivos Excel .xlsx.', status_code=400)

        try:
            wb = load_workbook(filename=BytesIO(file_obj.read()), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as exc:
            raise APIError('invalid_excel', 'El archivo no es un Excel válido.', status_code=400) from exc

        if len(rows) < 2:
            return Response({
                'success': False,
                'summary': {'total_rows': 0, 'valid_rows': 0, 'errors_count': 0},
                'errors': [{'row': 0, 'field': 'Archivo', 'message': 'El archivo no contiene filas de datos.'}],
                'preview': []
            })

        header_row = rows[0]
        # Parse headers mapping
        raw_mapping = {}
        for idx, cell in enumerate(header_row):
            if cell is not None:
                raw_mapping[_normalize_header(cell)] = idx

        # Translate headings to DB fields
        field_mapping = {}
        target_map = PROPERTY_HEADER_MAP if upload_type == 'properties' else USER_HEADER_MAP
        for k, v in raw_mapping.items():
            # Clean parentheses, question marks, asterisks, and underscores
            clean_k = re.sub(r'\(.*?\)', '', k)
            clean_k = clean_k.replace('?', '').replace('*', '')
            clean_k = clean_k.strip('_')
            if clean_k in target_map:
                field_mapping[target_map[clean_k]] = v

        # Validate required headers
        required_fields = ['direccion', 'propietario', 'tipo'] if upload_type == 'properties' else ['correo_electronico', 'nombres', 'apellidos', 'tipo_documento', 'numero_documento']
        missing_headers = [f for f in required_fields if f not in field_mapping]
        if missing_headers:
            return Response({
                'success': False,
                'summary': {'total_rows': 0, 'valid_rows': 0, 'errors_count': 0},
                'errors': [{'row': 1, 'field': 'Cabecera', 'message': f'Faltan columnas requeridas en el archivo: {", ".join(missing_headers)}'}],
                'preview': []
            })

        # DB Pre-fetching for fast validation
        database_addresses = {addr.lower() for addr in Property.objects.values_list('address', flat=True)}
        database_emails = {email.lower() for email in CustomUser.objects.values_list('email', flat=True)}
        database_docs = {doc for doc in CustomUser.objects.values_list('document_number', flat=True) if doc}

        # Keep track of file level uniqueness
        existing_addresses_in_file = set()
        existing_emails_in_file = set()
        existing_docs_in_file = set()

        errors = []
        preview = []

        for row_idx, row in enumerate(rows[1:], start=2):
            # Check empty row
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue

            # Construct row data
            row_data = {}
            for field, col_idx in field_mapping.items():
                if col_idx < len(row):
                    row_data[field] = _cell_str(row[col_idx])
                else:
                    row_data[field] = ''

            row_errors = []
            if upload_type == 'properties':
                # Properties validation logic
                address = row_data.get('direccion', '').strip()
                owner = row_data.get('propietario', '').strip()
                raw_type = row_data.get('tipo', '').strip().upper()
                raw_status = row_data.get('estado', '').strip().upper()

                if not address:
                    row_errors.append({'field': 'Dirección', 'message': 'La dirección es obligatoria.'})
                else:
                    if address.lower() in existing_addresses_in_file:
                        row_errors.append({'field': 'Dirección', 'message': f'Dirección duplicada en el archivo: "{address}".'})
                    else:
                        existing_addresses_in_file.add(address.lower())

                    if address.lower() in database_addresses:
                        row_errors.append({'field': 'Dirección', 'message': 'Ya existe un inmueble en el sistema con esta dirección.'})

                if not owner:
                    row_errors.append({'field': 'Título', 'message': 'El título es obligatorio.'})

                if not raw_type:
                    row_errors.append({'field': 'Tipo', 'message': 'El tipo de inmueble es obligatorio.'})
                elif raw_type not in PROPERTY_TYPE_CHOICES_MAP:
                    row_errors.append({'field': 'Tipo', 'message': f'Tipo no reconocido: "{raw_type}". Valores permitidos: APARTAMENTO, CASA, LOCAL, BODEGA.'})

                if raw_status and raw_status not in PROPERTY_STATUS_CHOICES_MAP:
                    row_errors.append({'field': 'Estado', 'message': f'Estado no reconocido: "{raw_status}". Valores permitidos: DISPONIBLE, ARRENDADO, MANTENIMIENTO.'})

                # Validate floats
                for f_label, f_key in [('Precio', 'precio'), ('Valor Administración', 'valor_administracion')]:
                    val = row_data.get(f_key, '')
                    if val:
                        try:
                            float(val)
                        except ValueError:
                            row_errors.append({'field': f_label, 'message': f'El campo "{f_label}" debe ser un valor numérico.'})

                # Validate integers
                for f_label, f_key in [('Habitaciones', 'habitaciones'), ('Baños', 'banos'), ('Salas', 'salas'), ('Cocinas', 'cocinas'), ('Garajes', 'garajes')]:
                    val = row_data.get(f_key, '')
                    if val:
                        try:
                            int(float(val))
                        except ValueError:
                            row_errors.append({'field': f_label, 'message': f'El campo "{f_label}" debe ser un número entero.'})

                # Validate booleans
                for f_label, f_key in [('Es Comercial? (S/N)', 'es_comercial'), ('En Conjunto? (S/N)', 'en_conjunto'), ('Administración Incluida? (S/N)', 'administracion_incluida')]:
                    val = row_data.get(f_key, '').strip().upper()
                    if val and val not in ('S', 'N', 'SI', 'NO', 'TRUE', 'FALSE', '1', '0', 'Y'):
                        row_errors.append({'field': f_label, 'message': f'El campo "{f_label}" debe ser S (Sí) o N (No).'})

                if not row_errors:
                    preview.append({
                        'row': row_idx,
                        'direccion': address,
                        'propietario': owner,
                        'tipo': PROPERTY_TYPE_CHOICES_MAP.get(raw_type, Property.Type.APARTMENT),
                        'estado': PROPERTY_STATUS_CHOICES_MAP.get(raw_status, Property.Status.AVAILABLE),
                        'precio': float(row_data.get('precio')) if row_data.get('precio') else 0.0,
                    })

            else:
                # Users validation logic
                email = row_data.get('correo_electronico', '').strip().lower()
                first_name = row_data.get('nombres', '').strip()
                last_name = row_data.get('apellidos', '').strip()
                doc_type = row_data.get('tipo_documento', '').strip().upper()
                doc_number = row_data.get('numero_documento', '').strip()
                phone = row_data.get('telefono', '').strip()
                raw_role = row_data.get('rol', '').strip().upper()

                if not email:
                    row_errors.append({'field': 'Correo Electrónico', 'message': 'El correo electrónico es obligatorio.'})
                elif not re.match(r'^[^@]+@[^@]+\.[^@]+$', email):
                    row_errors.append({'field': 'Correo Electrónico', 'message': 'Formato de correo electrónico no válido.'})
                else:
                    if email in existing_emails_in_file:
                        row_errors.append({'field': 'Correo Electrónico', 'message': f'Correo duplicado en el archivo: "{email}".'})
                    else:
                        existing_emails_in_file.add(email)

                    if email in database_emails:
                        row_errors.append({'field': 'Correo Electrónico', 'message': 'Ya existe un usuario registrado con este correo.'})

                if not first_name:
                    row_errors.append({'field': 'Nombres', 'message': 'El nombre es obligatorio.'})
                elif not re.match(r'^[a-zA-ZáéíóúÁÉÍÓÚñÑüÜ\s\.]+$', first_name):
                    row_errors.append({'field': 'Nombres', 'message': 'El nombre solo debe contener letras.'})

                if not last_name:
                    row_errors.append({'field': 'Apellidos', 'message': 'El apellido es obligatorio.'})
                elif not re.match(r'^[a-zA-ZáéíóúÁÉÍÓÚñÑüÜ\s\.]+$', last_name):
                    row_errors.append({'field': 'Apellidos', 'message': 'El apellido solo debe contener letras.'})

                if not doc_type:
                    row_errors.append({'field': 'Tipo Documento', 'message': 'El tipo de documento es obligatorio.'})
                elif doc_type not in USER_DOC_TYPE_CHOICES_MAP:
                    row_errors.append({'field': 'Tipo Documento', 'message': f'Tipo de documento no válido: "{doc_type}". Opciones válidas: CC, CE, PASAPORTE, NIT.'})

                if not doc_number:
                    row_errors.append({'field': 'Número Documento', 'message': 'El número de documento es obligatorio.'})
                elif not re.match(r'^[a-zA-Z0-9\-]+$', doc_number):
                    row_errors.append({'field': 'Número Documento', 'message': 'El número de documento debe contener solo letras, números y guiones.'})
                else:
                    if doc_number in existing_docs_in_file:
                        row_errors.append({'field': 'Número Documento', 'message': f'Número de documento duplicado en el archivo: "{doc_number}".'})
                    else:
                        existing_docs_in_file.add(doc_number)

                    if doc_number in database_docs:
                        row_errors.append({'field': 'Número Documento', 'message': 'Ya existe un usuario registrado con este número de documento.'})

                if phone:
                    if not phone.isdigit() or len(phone) != 10:
                        row_errors.append({'field': 'Teléfono', 'message': 'El teléfono debe contener exactamente 10 dígitos numéricos.'})

                if raw_role and raw_role not in USER_ROLE_CHOICES_MAP:
                    row_errors.append({'field': 'Rol', 'message': f'Rol no reconocido: "{raw_role}". Valores permitidos: ARRENDATARIO, ASISTENTE, TECNICO, ADMINISTRADOR.'})

                if not row_errors:
                    preview.append({
                        'row': row_idx,
                        'email': email,
                        'nombre': f"{first_name} {last_name}",
                        'documento': f"{doc_type} {doc_number}",
                        'rol': USER_ROLE_CHOICES_MAP.get(raw_role, CustomUser.Role.TENANT),
                    })

            for err in row_errors:
                errors.append({
                    'row': row_idx,
                    'field': err['field'],
                    'message': err['message']
                })

        total_rows = len(rows) - 1
        errors_count = len(errors)
        valid_rows = total_rows - len(set(err['row'] for err in errors))

        return Response({
            'success': errors_count == 0,
            'summary': {
                'total_rows': total_rows,
                'valid_rows': valid_rows,
                'errors_count': errors_count
            },
            'errors': errors,
            'preview': preview
        })


class BulkUploadImportView(APIView):
    """Valida el archivo y, si no tiene errores, realiza el guardado dentro de una transacción atómica."""
    permission_classes = [IsAdmin]
    parser_classes = [MultiPartParser]

    def post(self, request):
        upload_type = request.query_params.get('type')
        if upload_type not in ('properties', 'users'):
            raise APIError('invalid_type', 'El parámetro "type" debe ser "properties" o "users".', status_code=400)

        file_obj = request.FILES.get('file')
        if not file_obj:
            raise APIError('file_required', 'Debe enviar el archivo en el campo "file".', status_code=400)

        if not file_obj.name.lower().endswith('.xlsx'):
            raise APIError('invalid_file_type', 'Solo se admiten archivos Excel .xlsx.', status_code=400)

        try:
            wb = load_workbook(filename=BytesIO(file_obj.read()), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as exc:
            raise APIError('invalid_excel', 'El archivo no es un Excel válido.', status_code=400) from exc

        if len(rows) < 2:
            raise APIError('empty_file', 'El archivo no contiene filas de datos.', status_code=400)

        header_row = rows[0]
        # Parse headers mapping
        raw_mapping = {}
        for idx, cell in enumerate(header_row):
            if cell is not None:
                raw_mapping[_normalize_header(cell)] = idx

        # Translate headings to DB fields
        field_mapping = {}
        target_map = PROPERTY_HEADER_MAP if upload_type == 'properties' else USER_HEADER_MAP
        for k, v in raw_mapping.items():
            # Clean parentheses, question marks, asterisks, and underscores
            clean_k = re.sub(r'\(.*?\)', '', k)
            clean_k = clean_k.replace('?', '').replace('*', '')
            clean_k = clean_k.strip('_')
            if clean_k in target_map:
                field_mapping[target_map[clean_k]] = v

        # DB Pre-fetching for fast validation
        database_addresses = {addr.lower() for addr in Property.objects.values_list('address', flat=True)}
        database_emails = {email.lower() for email in CustomUser.objects.values_list('email', flat=True)}
        database_docs = {doc for doc in CustomUser.objects.values_list('document_number', flat=True) if doc}

        # Keep track of file level uniqueness
        existing_addresses_in_file = set()
        existing_emails_in_file = set()
        existing_docs_in_file = set()

        errors = []
        rows_to_create = []

        for row_idx, row in enumerate(rows[1:], start=2):
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue

            row_data = {}
            for field, col_idx in field_mapping.items():
                if col_idx < len(row):
                    row_data[field] = _cell_str(row[col_idx])
                else:
                    row_data[field] = ''

            row_errors = []
            if upload_type == 'properties':
                address = row_data.get('direccion', '').strip()
                owner = row_data.get('propietario', '').strip()
                raw_type = row_data.get('tipo', '').strip().upper()
                raw_status = row_data.get('estado', '').strip().upper()

                if not address:
                    row_errors.append({'field': 'Dirección', 'message': 'La dirección es obligatoria.'})
                else:
                    if address.lower() in existing_addresses_in_file:
                        row_errors.append({'field': 'Dirección', 'message': f'Dirección duplicada en el archivo: "{address}".'})
                    else:
                        existing_addresses_in_file.add(address.lower())

                    if address.lower() in database_addresses:
                        row_errors.append({'field': 'Dirección', 'message': 'Ya existe un inmueble en el sistema con esta dirección.'})

                if not owner:
                    row_errors.append({'field': 'Título', 'message': 'El título es obligatorio.'})

                if not raw_type:
                    row_errors.append({'field': 'Tipo', 'message': 'El tipo de inmueble es obligatorio.'})
                elif raw_type not in PROPERTY_TYPE_CHOICES_MAP:
                    row_errors.append({'field': 'Tipo', 'message': f'Tipo no reconocido: "{raw_type}".'})

                if raw_status and raw_status not in PROPERTY_STATUS_CHOICES_MAP:
                    row_errors.append({'field': 'Estado', 'message': f'Estado no reconocido: "{raw_status}".'})

                # floats
                price = 0.0
                admin_value = None
                for f_label, f_key in [('Precio', 'precio'), ('Valor Administración', 'valor_administracion')]:
                    val = row_data.get(f_key, '')
                    if val:
                        try:
                            f_val = float(val)
                            if f_key == 'precio':
                                price = f_val
                            else:
                                admin_value = f_val
                        except ValueError:
                            row_errors.append({'field': f_label, 'message': f'El campo "{f_label}" debe ser un valor numérico.'})

                # integers
                integers = {}
                for f_label, f_key in [('Habitaciones', 'habitaciones'), ('Baños', 'banos'), ('Salas', 'salas'), ('Cocinas', 'cocinas'), ('Garajes', 'garajes')]:
                    val = row_data.get(f_key, '')
                    if val:
                        try:
                            integers[f_key] = int(float(val))
                        except ValueError:
                            row_errors.append({'field': f_label, 'message': f'El campo "{f_label}" debe ser un número entero.'})
                    else:
                        integers[f_key] = None

                # booleans
                booleans = {}
                for f_label, f_key in [('Es Comercial? (S/N)', 'es_comercial'), ('En Conjunto? (S/N)', 'en_conjunto'), ('Administración Incluida? (S/N)', 'administracion_incluida')]:
                    val = row_data.get(f_key, '').strip().upper()
                    if val:
                        if val in ('S', 'SI', 'TRUE', '1', 'Y'):
                            booleans[f_key] = True
                        elif val in ('N', 'NO', 'FALSE', '0'):
                            booleans[f_key] = False
                        else:
                            row_errors.append({'field': f_label, 'message': f'El campo "{f_label}" debe ser S o N.'})
                    else:
                        booleans[f_key] = False

                if not row_errors:
                    rows_to_create.append({
                        'address': address,
                        'owner_name': owner,
                        'type': PROPERTY_TYPE_CHOICES_MAP.get(raw_type, Property.Type.APARTMENT),
                        'status': PROPERTY_STATUS_CHOICES_MAP.get(raw_status, Property.Status.AVAILABLE),
                        'city': row_data.get('ciudad', ''),
                        'building_name': row_data.get('edificio_conjunto', ''),
                        'unit_label': row_data.get('unidad_apartamento', ''),
                        'price': price,
                        'rooms': integers['habitaciones'],
                        'bathrooms': integers['banos'],
                        'living_rooms': integers['salas'],
                        'kitchens': integers['cocinas'],
                        'garages': integers['garajes'],
                        'is_commercial': booleans['es_comercial'],
                        'in_complex': booleans['en_conjunto'],
                        'admin_included': booleans['administracion_incluida'],
                        'admin_value': admin_value,
                        'google_maps_link': row_data.get('enlace_google_maps', ''),
                        'description': row_data.get('descripcion', ''),
                        'observations': row_data.get('observaciones', ''),
                    })
            else:
                email = row_data.get('correo_electronico', '').strip().lower()
                first_name = row_data.get('nombres', '').strip()
                last_name = row_data.get('apellidos', '').strip()
                doc_type = row_data.get('tipo_documento', '').strip().upper()
                doc_number = row_data.get('numero_documento', '').strip()
                phone = row_data.get('telefono', '').strip()
                raw_role = row_data.get('rol', '').strip().upper()

                if not email:
                    row_errors.append({'field': 'Correo Electrónico', 'message': 'El correo electrónico es obligatorio.'})
                elif not re.match(r'^[^@]+@[^@]+\.[^@]+$', email):
                    row_errors.append({'field': 'Correo Electrónico', 'message': 'Formato de correo electrónico no válido.'})
                else:
                    if email in existing_emails_in_file:
                        row_errors.append({'field': 'Correo Electrónico', 'message': f'Correo duplicado en el archivo: "{email}".'})
                    else:
                        existing_emails_in_file.add(email)

                    if email in database_emails:
                        row_errors.append({'field': 'Correo Electrónico', 'message': 'Ya existe un usuario registrado con este correo.'})

                if not first_name:
                    row_errors.append({'field': 'Nombres', 'message': 'El nombre es obligatorio.'})
                elif not re.match(r'^[a-zA-ZáéíóúÁÉÍÓÚñÑüÜ\s\.]+$', first_name):
                    row_errors.append({'field': 'Nombres', 'message': 'El nombre solo debe contener letras.'})

                if not last_name:
                    row_errors.append({'field': 'Apellidos', 'message': 'El apellido es obligatorio.'})
                elif not re.match(r'^[a-zA-ZáéíóúÁÉÍÓÚñÑüÜ\s\.]+$', last_name):
                    row_errors.append({'field': 'Apellidos', 'message': 'El apellido solo debe contener letras.'})

                if not doc_type:
                    row_errors.append({'field': 'Tipo Documento', 'message': 'El tipo de documento es obligatorio.'})
                elif doc_type not in USER_DOC_TYPE_CHOICES_MAP:
                    row_errors.append({'field': 'Tipo Documento', 'message': f'Tipo de documento no válido: "{doc_type}".'})

                if not doc_number:
                    row_errors.append({'field': 'Número Documento', 'message': 'El número de documento es obligatorio.'})
                elif not re.match(r'^[a-zA-Z0-9\-]+$', doc_number):
                    row_errors.append({'field': 'Número Documento', 'message': 'El número de documento debe contener solo letras, números y guiones.'})
                else:
                    if doc_number in existing_docs_in_file:
                        row_errors.append({'field': 'Número Documento', 'message': f'Número de documento duplicado en el archivo: "{doc_number}".'})
                    else:
                        existing_docs_in_file.add(doc_number)

                    if doc_number in database_docs:
                        row_errors.append({'field': 'Número Documento', 'message': 'Ya existe un usuario registrado con este número de documento.'})

                if phone:
                    if not phone.isdigit() or len(phone) != 10:
                        row_errors.append({'field': 'Teléfono', 'message': 'El teléfono debe contener exactamente 10 dígitos numéricos.'})

                if raw_role and raw_role not in USER_ROLE_CHOICES_MAP:
                    row_errors.append({'field': 'Rol', 'message': f'Rol no reconocido: "{raw_role}".'})

                if not row_errors:
                    rows_to_create.append({
                        'email': email,
                        'first_name': first_name,
                        'last_name': last_name,
                        'document_type': USER_DOC_TYPE_CHOICES_MAP.get(doc_type, CustomUser.DocumentType.CC),
                        'document_number': doc_number,
                        'phone': phone,
                        'role': USER_ROLE_CHOICES_MAP.get(raw_role, CustomUser.Role.TENANT),
                    })

            for err in row_errors:
                errors.append({
                    'row': row_idx,
                    'field': err['field'],
                    'message': err['message']
                })

        if errors:
            # If there are any errors, refuse the import
            return Response({
                'success': False,
                'message': 'No se pudo realizar la importación debido a que se encontraron errores en el archivo.',
                'errors': errors
            }, status=status.HTTP_400_BAD_REQUEST)

        # Atomic Transaction
        created_count = 0
        try:
            with transaction.atomic():
                if upload_type == 'properties':
                    for p_data in rows_to_create:
                        prop_status = p_data.pop('status', Property.Status.AVAILABLE)
                        prop = property_service.crear_propiedad(
                            created_by=request.user,
                            **p_data
                        )
                        if prop_status != Property.Status.AVAILABLE:
                            property_service.actualizar_propiedad(
                                prop,
                                updated_by=request.user,
                                status=prop_status
                            )
                        created_count += 1
                else:
                    for u_data in rows_to_create:
                        doc_num = u_data['document_number']
                        # Create user (role defined, temp password = doc_number, password_changed = False)
                        CustomUser.objects.create_user(
                            email=u_data['email'],
                            password=doc_num,
                            role=u_data['role'],
                            first_name=u_data['first_name'],
                            last_name=u_data['last_name'],
                            document_type=u_data['document_type'],
                            document_number=doc_num,
                            phone=u_data['phone'],
                            password_changed=False
                        )
                        created_count += 1
        except Exception as exc:
            raise APIError('import_failed', f'La importación falló debido a un error inesperado al escribir en la base de datos: {str(exc)}', status_code=500)

        return Response({
            'success': True,
            'message': f'Se importaron exitosamente {created_count} registros.',
            'created_count': created_count
        }, status=status.HTTP_201_CREATED)
