from io import BytesIO
from django.test import TestCase
from rest_framework import status
from rest_framework.test import APIClient
from openpyxl import Workbook

from pot.models import CustomUser, Property


class BulkUploadTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.password = 'AdminPass123!'
        
        # Create an ADMIN user
        self.admin = CustomUser.objects.create_user(
            email='admin@test.com',
            password=self.password,
            role=CustomUser.Role.ADMIN,
            document_number='111222333',
            document_type=CustomUser.DocumentType.CC,
            password_changed=True,
        )

        # Create a TENANT user for permission testing
        self.tenant = CustomUser.objects.create_user(
            email='tenant@test.com',
            password=self.password,
            role=CustomUser.Role.TENANT,
            document_number='444555666',
            document_type=CustomUser.DocumentType.CC,
            password_changed=True,
        )

    def test_permission_denied_for_tenants(self):
        # Authenticate as a tenant
        self.client.force_authenticate(user=self.tenant)
        
        r1 = self.client.get('/api/v1/admin/bulk-upload/template/?type=properties')
        self.assertEqual(r1.status_code, status.HTTP_403_FORBIDDEN)

        r2 = self.client.get('/api/v1/admin/bulk-upload/export/?type=users')
        self.assertEqual(r2.status_code, status.HTTP_403_FORBIDDEN)

        r3 = self.client.post('/api/v1/admin/bulk-upload/validate/?type=properties')
        self.assertEqual(r3.status_code, status.HTTP_403_FORBIDDEN)

    def test_download_templates_success(self):
        self.client.force_authenticate(user=self.admin)
        
        # Properties template
        r = self.client.get('/api/v1/admin/bulk-upload/template/?type=properties')
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertEqual(r['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertIn('attachment', r['Content-Disposition'])
        
        # Users template
        r = self.client.get('/api/v1/admin/bulk-upload/template/?type=users')
        self.assertEqual(r.status_code, status.HTTP_200_OK)

    def test_export_success(self):
        # Create some dummy properties and users
        Property.objects.create(
            code='PRO-99991',
            address='Calle Falsa 123',
            type=Property.Type.HOUSE,
            owner_name='Homero Simpson'
        )
        
        self.client.force_authenticate(user=self.admin)
        
        # Export properties
        r = self.client.get('/api/v1/admin/bulk-upload/export/?type=properties')
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertEqual(r['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Export users
        r = self.client.get('/api/v1/admin/bulk-upload/export/?type=users')
        self.assertEqual(r.status_code, status.HTTP_200_OK)

    def test_validation_and_import_properties(self):
        self.client.force_authenticate(user=self.admin)
        
        # 1. Create an in-memory workbook with valid data
        wb = Workbook()
        ws = wb.active
        ws.append(["Dirección*", "Título*", "Tipo*", "Ciudad", "Precio"])
        ws.append(["Carrera 15 # 85-30", "Alicia Gómez", "APARTAMENTO", "Bogotá", "1200000"])
        ws.append(["Autopista Norte # 100-22", "Carlos Díaz", "LOCAL", "Medellín", "3500000"])
        
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        out.name = 'test_properties.xlsx'

        # 2. Test validate
        r_val = self.client.post(
            '/api/v1/admin/bulk-upload/validate/?type=properties',
            {'file': out},
            format='multipart'
        )
        self.assertEqual(r_val.status_code, status.HTTP_200_OK)
        self.assertTrue(r_val.data['success'])
        self.assertEqual(r_val.data['summary']['total_rows'], 2)
        self.assertEqual(r_val.data['summary']['errors_count'], 0)

        # Reset pointer
        out.seek(0)

        # 3. Test import
        r_imp = self.client.post(
            '/api/v1/admin/bulk-upload/import/?type=properties',
            {'file': out},
            format='multipart'
        )
        self.assertEqual(r_imp.status_code, status.HTTP_201_CREATED)
        self.assertTrue(r_imp.data['success'])
        self.assertEqual(r_imp.data['created_count'], 2)

        # Verify DB entries exist
        self.assertTrue(Property.objects.filter(address='Carrera 15 # 85-30').exists())
        self.assertTrue(Property.objects.filter(address='Autopista Norte # 100-22').exists())

    def test_validation_and_import_users(self):
        self.client.force_authenticate(user=self.admin)
        
        # 1. Create workbook with users (valid and invalid)
        wb = Workbook()
        ws = wb.active
        ws.append(["Correo Electrónico*", "Nombres*", "Apellidos*", "Tipo Documento*", "Número Documento*", "Teléfono", "Rol"])
        # Row 2: Valid User
        ws.append(["jorge.val@test.com", "Jorge", "Valenzuela", "CC", "999888777", "3109876543", "ARRENDATARIO"])
        # Row 3: Invalid Email
        ws.append(["correo-invalido", "Felipe", "Rojas", "CE", "111222444", "3201234567", "TECNICO"])
        
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        out.name = 'test_users.xlsx'

        # 2. Validate must return success=False and show 1 error
        r_val = self.client.post(
            '/api/v1/admin/bulk-upload/validate/?type=users',
            {'file': out},
            format='multipart'
        )
        self.assertEqual(r_val.status_code, status.HTTP_200_OK)
        self.assertFalse(r_val.data['success'])
        self.assertEqual(r_val.data['summary']['errors_count'], 1)
        self.assertEqual(r_val.data['errors'][0]['row'], 3)
        self.assertEqual(r_val.data['errors'][0]['field'], 'Correo Electrónico')

        # Reset pointer
        out.seek(0)

        # 3. Import must fail completely (all-or-nothing check)
        r_imp = self.client.post(
            '/api/v1/admin/bulk-upload/import/?type=users',
            {'file': out},
            format='multipart'
        )
        self.assertEqual(r_imp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(r_imp.data['success'])
        
        # Verify that EVEN the valid user was NOT created due to database rollback / error block
        self.assertFalse(CustomUser.objects.filter(email='jorge.val@test.com').exists())

    def test_all_or_nothing_atomic_transaction(self):
        self.client.force_authenticate(user=self.admin)
        
        # 1. Create a spreadsheet with one duplicate document number (which exists in DB: self.admin has '111222333')
        wb = Workbook()
        ws = wb.active
        ws.append(["Correo Electrónico*", "Nombres*", "Apellidos*", "Tipo Documento*", "Número Documento*", "Teléfono", "Rol"])
        # Valid user 1
        ws.append(["maria@test.com", "Maria", "Silva", "CC", "777666555", "3003003030", "ARRENDATARIO"])
        # Duplicate document number of admin (fails DB unique check)
        ws.append(["maria2@test.com", "Maria Dos", "Silva", "CC", "111222333", "3003003030", "ARRENDATARIO"])

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        out.name = 'test_atomic.xlsx'

        # 2. Try to import. Should fail.
        r_imp = self.client.post(
            '/api/v1/admin/bulk-upload/import/?type=users',
            {'file': out},
            format='multipart'
        )
        self.assertEqual(r_imp.status_code, status.HTTP_400_BAD_REQUEST)
        
        # 3. Ensure 'maria@test.com' was NOT created because transaction rolled back or import blocked
        self.assertFalse(CustomUser.objects.filter(email='maria@test.com').exists())
