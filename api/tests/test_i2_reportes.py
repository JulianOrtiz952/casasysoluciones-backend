"""Pruebas i2 reportes — HU-09 (CP-RF-29, CP-RF-30)."""

import io
from datetime import timedelta

from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from pot.models import CustomUser, Property, Ticket, UserPropertyAssociation


def _open_ticket(tenant, prop, **kwargs):
    defaults = {
        'property': prop,
        'tenant': tenant,
        'description': 'Daño reportado en baño principal.',
        'damage_type': Ticket.DamageType.PLUMBING,
        'priority': Ticket.Priority.HIGH,
        'status': Ticket.Status.OPEN,
        'title': 'Fuga baño',
    }
    defaults.update(kwargs)
    return Ticket.objects.create(**defaults)


class ReportsAPITests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = CustomUser.objects.create_user(
            email='admin-rep@test.com',
            password='AdminPass123!',
            role=CustomUser.Role.ADMIN,
            password_changed=True,
        )
        self.assistant = CustomUser.objects.create_user(
            email='assistant-rep@test.com',
            password='AssistPass123!',
            role=CustomUser.Role.ASSISTANT,
            password_changed=True,
        )
        self.tenant = CustomUser.objects.create_user(
            email='tenant-rep@test.com',
            password='TenantPass123!',
            role=CustomUser.Role.TENANT,
            password_changed=True,
            first_name='Ana',
            last_name='Arrendatario',
        )
        self.property = Property.objects.create(
            code='PRO-REP-01',
            address='Calle Reportes 1',
            type=Property.Type.APARTMENT,
            owner_name='Dueño',
            status=Property.Status.RENTED,
        )
        UserPropertyAssociation.objects.create(user=self.tenant, property=self.property)

    def _admin(self):
        self.client.force_authenticate(user=self.admin)

    def test_cp_rf_30_assistant_forbidden_on_reports_rf30(self):
        """CP-RF-30: reportes administrativos solo ADMIN (assistant excluido)."""
        self.client.force_authenticate(user=self.assistant)
        endpoints = [
            '/api/v1/reports/ticket-traffic-light/',
            '/api/v1/reports/summary/',
            '/api/v1/reports/properties-with-open-tickets/',
            '/api/v1/reports/tenants-with-active-tickets/',
            f'/api/v1/reports/properties/{self.property.id}/repair-history/',
            '/api/v1/reports/export/excel/',
            '/api/v1/search/?q=PRO',
        ]
        for url in endpoints:
            r = self.client.get(url)
            self.assertEqual(r.status_code, status.HTTP_403_FORBIDDEN, url)

    def test_cp_rf_29_reports_traffic_light_pending_resolution_rf29(self):
        """CP-RF-29: semáforo consolidado y pending_resolution en reportes."""
        _open_ticket(self.tenant, self.property, priority=Ticket.Priority.HIGH)
        _open_ticket(
            self.tenant,
            self.property,
            priority=Ticket.Priority.MEDIUM,
            status=Ticket.Status.ACCEPTED,
        )
        self._admin()
        r = self.client.get('/api/v1/reports/ticket-traffic-light/')
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertEqual(r.data['pending_resolution'], 2)
        self.assertIn('traffic_light', r.data)
        self.assertEqual(
            sum(r.data['traffic_light'].values()),
            r.data['pending_resolution'],
        )

    def test_traffic_light_filtered_by_property(self):
        other_prop = Property.objects.create(
            code='PRO-REP-02',
            address='Calle Reportes 2',
            type=Property.Type.HOUSE,
            owner_name='Dueño 2',
            status=Property.Status.RENTED,
        )
        _open_ticket(self.tenant, self.property)
        _open_ticket(self.tenant, other_prop)
        self._admin()
        r = self.client.get(f'/api/v1/reports/ticket-traffic-light/?property_id={self.property.id}')
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertEqual(r.data['pending_resolution'], 1)

    def test_cp_rf_30_properties_with_open_tickets_rf30(self):
        """CP-RF-30: inmuebles arrendados con tickets abiertos."""
        _open_ticket(self.tenant, self.property)
        Property.objects.create(
            code='PRO-REP-AVAIL',
            address='Calle Disponible',
            type=Property.Type.LOCAL,
            owner_name='Dueño',
            status=Property.Status.AVAILABLE,
        )
        self._admin()
        r = self.client.get('/api/v1/reports/properties-with-open-tickets/')
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertEqual(len(r.data), 1)
        self.assertEqual(r.data[0]['property_id'], self.property.id)
        self.assertEqual(r.data[0]['open_tickets_count'], 1)
        self.assertEqual(r.data[0]['tenant']['email'], self.tenant.email)

    def test_cp_rf_30_tenants_with_active_tickets_rf30(self):
        """CP-RF-30: inquilinos con tickets activos."""
        _open_ticket(self.tenant, self.property, status=Ticket.Status.IN_PROGRESS)
        self._admin()
        r = self.client.get('/api/v1/reports/tenants-with-active-tickets/')
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertEqual(len(r.data), 1)
        self.assertEqual(r.data[0]['tenant_id'], self.tenant.id)
        self.assertEqual(r.data[0]['active_tickets_count'], 1)

    def test_cp_rf_30_repair_history_closed_tickets_rf30(self):
        """CP-RF-30: historial de reparaciones del inmueble."""
        closed = _open_ticket(self.tenant, self.property, status=Ticket.Status.CLOSED)
        closed.assigned_contractor_name = 'Maestro López'
        closed.save(update_fields=['assigned_contractor_name', 'updated_at'])
        _open_ticket(self.tenant, self.property, status=Ticket.Status.OPEN)
        self._admin()
        r = self.client.get(f'/api/v1/reports/properties/{self.property.id}/repair-history/')
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertEqual(r.data['repairs_count'], 1)
        self.assertEqual(r.data['repairs'][0]['public_code'], closed.public_code)
        self.assertEqual(r.data['repairs'][0]['assigned_contractor_name'], 'Maestro López')

    def test_summary_by_status_priority_damage(self):
        _open_ticket(self.tenant, self.property, damage_type=Ticket.DamageType.ELECTRICITY)
        _open_ticket(
            self.tenant,
            self.property,
            status=Ticket.Status.CLOSED,
            damage_type=Ticket.DamageType.PLUMBING,
        )
        self._admin()
        r = self.client.get('/api/v1/reports/summary/')
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertEqual(r.data['total_tickets'], 2)
        self.assertIn('by_status', r.data)
        self.assertIn('by_priority', r.data)
        self.assertIn('by_damage_type', r.data)

    def test_cp_rf_30_export_excel_with_filters_rf30(self):
        """CP-RF-30: exportación Excel con filtros date_from y property_id."""
        old = _open_ticket(self.tenant, self.property)
        Ticket.objects.filter(pk=old.pk).update(
            created_at=timezone.now() - timedelta(days=40),
        )
        recent = _open_ticket(self.tenant, self.property, public_code=None)
        date_from = (timezone.now() - timedelta(days=7)).date().isoformat()
        self._admin()
        r = self.client.get(
            f'/api/v1/reports/export/excel/?property_id={self.property.id}&date_from={date_from}',
        )
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertIn('spreadsheetml', r['Content-Type'])
        wb = load_workbook(io.BytesIO(r.content))
        rows = list(wb.active.iter_rows(values_only=True))
        self.assertGreaterEqual(len(rows), 2)
        codes = [row[0] for row in rows[1:]]
        self.assertIn(recent.public_code, codes)
        self.assertNotIn(old.public_code, codes)

    def test_global_search(self):
        self._admin()
        r = self.client.get(f'/api/v1/search/?q={self.property.code}')
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertTrue(r.data['properties'])
        self.assertEqual(r.data['properties'][0]['code'], self.property.code)

    def test_tenant_forbidden_on_reports(self):
        self.client.force_authenticate(user=self.tenant)
        r = self.client.get('/api/v1/reports/summary/')
        self.assertEqual(r.status_code, status.HTTP_403_FORBIDDEN)
