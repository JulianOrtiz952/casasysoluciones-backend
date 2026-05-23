"""Carga inicial de inmuebles y arrendatarios desde Excel (compromiso Acta 2)."""

from __future__ import annotations

import re
from io import BytesIO

from django.db import transaction
from openpyxl import load_workbook

from pot.models import CustomUser, Property
from pot.services import property_service, user_service
from pot.services.user_service import UserServiceError


class ImportExcelError(Exception):
    def __init__(self, code, message, details=None):
        self.code = code
        self.message = message
        self.details = details or {}
        super().__init__(message)


# Encabezados aceptados (minúsculas, sin acentos opcionales vía normalización)
HEADER_ALIASES = {
    'direccion': 'address',
    'address': 'address',
    'tipo': 'type',
    'type': 'type',
    'ciudad': 'city',
    'city': 'city',
    'edificio': 'building_name',
    'conjunto': 'building_name',
    'building_name': 'building_name',
    'unidad': 'unit_label',
    'apartamento': 'unit_label',
    'unit_label': 'unit_label',
    'propietario': 'owner_name',
    'dueno': 'owner_name',
    'owner_name': 'owner_name',
    'estado': 'status',
    'status': 'status',
    'email': 'email',
    'correo': 'email',
    'documento': 'document_number',
    'cedula': 'document_number',
    'document_number': 'document_number',
    'tipo_documento': 'document_type',
    'document_type': 'document_type',
    'nombre': 'first_name',
    'first_name': 'first_name',
    'apellido': 'last_name',
    'last_name': 'last_name',
    'telefono': 'phone',
    'phone': 'phone',
}

PROPERTY_TYPE_MAP = {
    'APARTMENT': Property.Type.APARTMENT,
    'APARTAMENTO': Property.Type.APARTMENT,
    'APT': Property.Type.APARTMENT,
    'HOUSE': Property.Type.HOUSE,
    'CASA': Property.Type.HOUSE,
    'LOCAL': Property.Type.LOCAL,
    'LOCAL COMERCIAL': Property.Type.LOCAL,
    'WAREHOUSE': Property.Type.WAREHOUSE,
    'BODEGA': Property.Type.WAREHOUSE,
}

PROPERTY_STATUS_MAP = {
    'AVAILABLE': Property.Status.AVAILABLE,
    'DISPONIBLE': Property.Status.AVAILABLE,
    'RENTED': Property.Status.RENTED,
    'ARRENDADO': Property.Status.RENTED,
    'MAINTENANCE': Property.Status.MAINTENANCE,
    'MANTENIMIENTO': Property.Status.MAINTENANCE,
}

DOCUMENT_TYPE_MAP = {
    'CC': CustomUser.DocumentType.CC,
    'CEDULA': CustomUser.DocumentType.CC,
    'CE': CustomUser.DocumentType.CE,
    'PASAPORTE': CustomUser.DocumentType.PASSPORT,
    'PASSPORT': CustomUser.DocumentType.PASSPORT,
    'NIT': CustomUser.DocumentType.NIT,
}


def _normalize_header(value):
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = re.sub(r'\s+', '_', text)
    text = (
        text.replace('á', 'a')
        .replace('é', 'e')
        .replace('í', 'i')
        .replace('ó', 'o')
        .replace('ú', 'u')
        .replace('ñ', 'n')
    )
    return text


def _cell_str(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _map_property_type(raw):
    if not raw:
        return Property.Type.APARTMENT
    key = _cell_str(raw).upper()
    if key not in PROPERTY_TYPE_MAP:
        raise ImportExcelError(
            'invalid_property_type',
            f'Tipo de inmueble no reconocido: {raw}',
            {'value': raw},
        )
    return PROPERTY_TYPE_MAP[key]


def _map_property_status(raw):
    if not raw:
        return None
    key = _cell_str(raw).upper()
    if key not in PROPERTY_STATUS_MAP:
        raise ImportExcelError(
            'invalid_property_status',
            f'Estado de inmueble no reconocido: {raw}',
            {'value': raw},
        )
    return PROPERTY_STATUS_MAP[key]


def _map_document_type(raw):
    if not raw:
        return CustomUser.DocumentType.CC
    key = _cell_str(raw).upper()
    return DOCUMENT_TYPE_MAP.get(key, CustomUser.DocumentType.CC)


def _parse_workbook(file_bytes):
    try:
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ImportExcelError('invalid_file', 'El archivo no es un Excel válido (.xlsx).') from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        raise ImportExcelError('empty_file', 'El archivo no contiene filas de datos.')
    return rows


def _build_column_map(header_row):
    column_map = {}
    for idx, cell in enumerate(header_row):
        normalized = _normalize_header(cell)
        field = HEADER_ALIASES.get(normalized)
        if field:
            column_map[field] = idx
    if 'address' not in column_map or 'owner_name' not in column_map:
        raise ImportExcelError(
            'missing_columns',
            'Faltan columnas obligatorias: direccion y propietario.',
            {'required': ['direccion', 'propietario']},
        )
    return column_map


def _row_dict(row, column_map):
    data = {}
    for field, idx in column_map.items():
        if idx < len(row):
            data[field] = _cell_str(row[idx])
    return data


def importar_desde_excel(file_bytes, created_by, *, send_credentials=False, request=None, dry_run=False):
    """
    Importa filas del Excel: crea inmuebles y, si hay email, arrendatarios asociados.

    Retorna resumen con contadores y errores por fila.
    """
    rows = _parse_workbook(file_bytes)
    column_map = _build_column_map(rows[0])

    summary = {
        'dry_run': dry_run,
        'rows_total': 0,
        'properties_created': 0,
        'tenants_created': 0,
        'associations_created': 0,
        'rows_skipped': 0,
        'errors': [],
    }

    for row_number, row in enumerate(rows[1:], start=2):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        summary['rows_total'] += 1
        try:
            _process_row(
                _row_dict(row, column_map),
                created_by,
                summary,
                send_credentials=send_credentials,
                request=request,
                dry_run=dry_run,
                row_number=row_number,
            )
        except (ImportExcelError, user_service.UserServiceError, property_service.PropertyServiceError) as exc:
            code = getattr(exc, 'code', 'row_error')
            message = getattr(exc, 'message', str(exc))
            details = getattr(exc, 'details', {})
            summary['errors'].append(
                {
                    'row': row_number,
                    'code': code,
                    'message': message,
                    'details': details,
                }
            )

    return summary


def _process_row(data, created_by, summary, *, send_credentials, request, dry_run, row_number):
    address = data.get('address', '').strip()
    owner_name = data.get('owner_name', '').strip()
    if not address:
        summary['rows_skipped'] += 1
        return
    if not owner_name:
        raise ImportExcelError('owner_required', 'El propietario es obligatorio.', {'row': row_number})

    prop_type = _map_property_type(data.get('type'))
    prop_status = _map_property_status(data.get('status'))
    email = data.get('email', '').strip().lower()

    if dry_run:
        summary['properties_created'] += 1
        if email:
            summary['tenants_created'] += int(not CustomUser.objects.filter(email__iexact=email).exists())
            summary['associations_created'] += 1
        return

    with transaction.atomic():
        existing_prop = Property.objects.filter(address__iexact=address).first()
        if existing_prop:
            prop = existing_prop
        else:
            prop = property_service.crear_propiedad(
                created_by,
                address=address,
                type=prop_type,
                owner_name=owner_name,
                city=data.get('city', ''),
                building_name=data.get('building_name', ''),
                unit_label=data.get('unit_label', ''),
            )
            summary['properties_created'] += 1

        if prop_status and prop.status != prop_status:
            property_service.actualizar_propiedad(prop, created_by, status=prop_status)

        if not email:
            return

        tenant_fields = {
            'first_name': data.get('first_name', ''),
            'last_name': data.get('last_name', ''),
            'phone': data.get('phone', ''),
            'document_type': _map_document_type(data.get('document_type')),
            'document_number': data.get('document_number') or None,
        }

        existing = CustomUser.objects.filter(email__iexact=email).first()
        if existing:
            if existing.role != CustomUser.Role.TENANT:
                raise ImportExcelError(
                    'not_tenant',
                    f'El correo {email} pertenece a un usuario que no es arrendatario.',
                )
            user_service.asociar_inmueble_arrendatario(
                existing,
                prop,
                created_by,
                request=request,
                notify=send_credentials,
            )
        else:
            if not tenant_fields.get('document_number'):
                raise ImportExcelError(
                    'document_required',
                    'Para crear arrendatario se requiere documento (cedula).',
                    {'row': row_number},
                )
            user_service.crear_arrendatario(
                created_by,
                email=email,
                property_ids=[prop.pk],
                request=request,
                send_credentials=send_credentials,
                **tenant_fields,
            )
            summary['tenants_created'] += 1
            summary['associations_created'] += 1
            return

        summary['associations_created'] += 1
